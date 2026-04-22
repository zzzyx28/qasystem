import pandas as pd
import openpyxl
from typing import Dict, List, Any, Tuple
import numpy as np
from datetime import datetime, date
import logging
import os
import csv
from . import BaseProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    def __init__(self):
        # 更新支持的文件格式
        self.supported_formats = ['.xlsx', '.xls', '.csv']

    def _convert_cell_value(self, value):
        """转换单元格值，确保适合MySQL插入"""
        if pd.isna(value):  # 处理NaN值
            return ''
        # 为了尽量保留 Excel 原始显示内容，默认将单元格值转换为字符串而不是对日期做统一格式化
        try:
            return str(value)
        except Exception:
            return ''

    def process_excel(self, file_path: str) -> Dict[str, Any]:
        """
        处理Excel或CSV文件，返回结构化数据
        返回格式: {
            'file_name': str,
            'sheets': [
                {
                    'sheet_name': str,
                    'data': List[List],  # 原始数据
                    'columns': List[str],  # 列名
                    'metadata': Dict
                }
            ]
        }
        """
        # 获取文件扩展名
        _, file_ext = os.path.splitext(file_path)
        file_ext = file_ext.lower()

        # 根据文件类型选择处理方法
        if file_ext == '.csv':
            return self._process_csv(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self._process_excel_file(file_path)
        else:
            logger.error(f"不支持的文件格式: {file_ext}")
            return {'error': f'不支持的文件格式: {file_ext}'}
        # 获取文件扩展名
        _, file_ext = os.path.splitext(file_path)
        file_ext = file_ext.lower()

        # 根据文件类型选择处理方法
        if file_ext == '.csv':
            return self._process_csv(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self._process_excel_file(file_path)
        else:
            logger.error(f"不支持的文件格式: {file_ext}")
            return {'error': f'不支持的文件格式: {file_ext}'}

    def _process_excel_file(self, file_path: str) -> Dict[str, Any]:
        """处理Excel文件 (.xlsx, .xls)"""
        try:
            # 使用pandas读取Excel
            excel_data = pd.ExcelFile(file_path)
            result = {
                'file_name': os.path.basename(file_path),
                'sheets': [],
                'total_sheets': len(excel_data.sheet_names),
                'processed_at': datetime.now().isoformat(),
                'file_type': 'excel'
            }

            for sheet_name in excel_data.sheet_names:
                # 读取每个sheet，强制按字符串读取以保留原始单元格文本（避免日期被统一格式化）
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, keep_default_na=False)
                except Exception:
                    # 回退到默认读取方式
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                # 确保所有单元格都为字符串，缺失值为空字符串
                try:
                    df = df.astype(str).fillna('')
                except Exception:
                    pass

                # 处理当前sheet
                sheet_info = self._process_dataframe(df, sheet_name)
                result['sheets'].append(sheet_info)

            return result

        except Exception as e:
            logger.error(f"使用pandas处理Excel失败: {e}")
            return self._fallback_excel_process(file_path)

    def _process_csv(self, file_path: str) -> Dict[str, Any]:
        """处理CSV文件"""
        try:
            with open(file_path, 'rb') as f:
                header = f.read(8)
            if header.startswith(b'PK\x03\x04') or header.startswith(b'\xD0\xCF\x11\xE0'):
                logger.warning("文件扩展名为CSV但内容为Excel二进制，自动按Excel解析: %s", file_path)
                return self._process_excel_file(file_path)

            # 尝试自动检测分隔符和编码
            result = {
                'file_name': os.path.basename(file_path),
                'sheets': [],
                'total_sheets': 1,  # CSV只有一个"sheet"
                'processed_at': datetime.now().isoformat(),
                'file_type': 'csv'
            }

            encodings = ['utf-8-sig', 'utf-8', 'gb18030', 'gbk', 'gb2312', 'cp936', 'big5', 'latin1']
            delimiters = [',', ';', '\t', '|']

            best_df = None
            best_score = -1

            for encoding in encodings:
                for delimiter in delimiters + [None]:
                    try:
                        read_kwargs = {
                            'encoding': encoding,
                            'on_bad_lines': 'skip'
                        }
                        if delimiter is None:
                            read_kwargs.update({'sep': None, 'engine': 'python'})
                        else:
                            read_kwargs.update({'sep': delimiter})

                        sample_df = pd.read_csv(file_path, nrows=100, **read_kwargs)
                        if sample_df.empty and sample_df.shape[1] == 0:
                            continue

                        score = sample_df.shape[1]
                        if delimiter == ',':
                            score += 0.2
                        if encoding in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk'):
                            score += 0.1

                        if score > best_score:
                            best_score = score
                            best_df = pd.read_csv(file_path, **read_kwargs)
                    except Exception:
                        continue

            if best_df is None:
                best_df = pd.read_csv(
                    file_path,
                    sep=None,
                    engine='python',
                    encoding='utf-8',
                    encoding_errors='replace',
                    on_bad_lines='skip'
                )

            df = best_df
            # 将所有列转换为字符串以尽量保留原始文本表示
            try:
                df = df.astype(str).fillna('')
            except Exception:
                pass

            # 处理CSV数据（当作单个sheet）
            sheet_name = os.path.basename(file_path).replace('.csv', '')
            sheet_info = self._process_dataframe(df, sheet_name)
            result['sheets'].append(sheet_info)

            return result

        except Exception as e:
            logger.error(f"处理CSV文件失败: {e}")
            return self._fallback_csv_process(file_path)

    def _process_dataframe(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """通用DataFrame处理逻辑"""
        # 转换列名为字符串
        columns = [str(col) for col in df.columns.tolist()]

        # 处理数据：转换每个单元格值
        data = []
        for _, row in df.iterrows():
            row_data = []
            for cell in row:
                row_data.append(self._convert_cell_value(cell))
            data.append(row_data)

        return {
            'sheet_name': sheet_name,
            'columns': columns,
            'data': data,
            'shape': df.shape,
            'metadata': {
                'row_count': df.shape[0],
                'col_count': df.shape[1],
                'data_types': {col: str(df[col].dtype) for col in df.columns}
            }
        }

    def _fallback_excel_process(self, file_path: str) -> Dict[str, Any]:
        """使用openpyxl作为Excel备选方案"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            result = {
                'file_name': os.path.basename(file_path),
                'sheets': [],
                'total_sheets': len(wb.sheetnames),
                'processed_at': datetime.now().isoformat(),
                'file_type': 'excel'
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []

                # 获取最大行和列
                max_row = ws.max_row
                max_col = ws.max_column

                # 读取数据
                for row in ws.iter_rows(min_row=1, max_row=max_row,
                                        min_col=1, max_col=max_col,
                                        values_only=True):
                    row_data = []
                    for cell in row:
                        row_data.append(self._convert_cell_value(cell))
                    data.append(row_data)

                if data:
                    columns = [str(cell) for cell in data[0]] if max_row > 0 else []
                    sheet_data = data[1:] if max_row > 1 else []
                else:
                    columns = []
                    sheet_data = []

                sheet_info = {
                    'sheet_name': sheet_name,
                    'columns': columns,
                    'data': sheet_data,
                    'shape': (len(sheet_data), len(columns)),
                    'metadata': {
                        'row_count': len(sheet_data),
                        'col_count': len(columns)
                    }
                }

                result['sheets'].append(sheet_info)

            return result

        except Exception as e:
            logger.error(f"使用openpyxl处理Excel失败: {e}")
            return {'error': str(e), 'file_name': os.path.basename(file_path)}

    def _fallback_csv_process(self, file_path: str) -> Dict[str, Any]:
        """使用Python内置csv模块作为CSV备选方案"""
        try:
            result = {
                'file_name': os.path.basename(file_path),
                'sheets': [],
                'total_sheets': 1,
                'processed_at': datetime.now().isoformat(),
                'file_type': 'csv'
            }

            data = []
            # 尝试不同编码
            encodings = ['utf-8-sig', 'utf-8', 'gb18030', 'gbk', 'gb2312', 'cp936', 'latin1']

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        # 尝试不同分隔符
                        dialect = csv.Sniffer().sniff(f.read(1024))
                        f.seek(0)
                        reader = csv.reader(f, dialect)
                        data = list(reader)
                        break
                except:
                    continue

            if not data:
                # 如果自动检测失败，使用默认逗号分隔符
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    reader = csv.reader(f)
                    data = list(reader)

            if data:
                columns = [str(cell) for cell in data[0]] if data else []
                sheet_data = data[1:] if len(data) > 1 else []
            else:
                columns = []
                sheet_data = []

            sheet_name = os.path.basename(file_path).replace('.csv', '')
            sheet_info = {
                'sheet_name': sheet_name,
                'columns': columns,
                'data': sheet_data,
                'shape': (len(sheet_data), len(columns)),
                'metadata': {
                    'row_count': len(sheet_data),
                    'col_count': len(columns)
                }
            }

            result['sheets'].append(sheet_info)
            return result

        except Exception as e:
            logger.error(f"使用csv模块处理CSV失败: {e}")
            return {'error': str(e), 'file_name': os.path.basename(file_path)}

    def generate_sql_schema(self, excel_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        为Excel数据生成SQL建表语句
        """
        schemas = []

        for sheet in excel_data['sheets']:
            sheet_name = sheet['sheet_name']
            columns = sheet['columns']

            # 清理表名和列名（移除特殊字符）
            clean_table_name = f"excel_{excel_data['file_name'].replace('.', '_')}_{sheet_name}"
            clean_table_name = ''.join(c for c in clean_table_name if c.isalnum() or c == '_')

            # 生成SQL
            sql_columns = []
            for col in columns:
                clean_col = ''.join(c for c in str(col) if c.isalnum() or c == '_')
                sql_columns.append(f"`{clean_col}` TEXT")

            create_table_sql = f"""
            CREATE TABLE IF NOT EXISTS `{clean_table_name}` (
                id INT AUTO_INCREMENT PRIMARY KEY,
                {', '.join(sql_columns)},
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """

            schemas.append({
                'table_name': clean_table_name,
                'sql': create_table_sql,
                'sheet_name': sheet_name
            })

        return schemas

    def process(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """
        处理Excel或CSV文件，返回包含 content 和 metadata 的字典
        """
        metadata = {
            'file_type': 'excel' if file_path.lower().endswith(('.xlsx', '.xls')) else 'csv',
            'processor': 'pandas',
            'supported_formats': ['.xlsx', '.xls', '.csv']
        }
        
        result = self.process_excel(file_path)
        if 'error' in result:
            metadata['error'] = result['error']
            return {
                'content': '',
                'metadata': metadata
            }
        
        return {
            'content': result,
            'metadata': metadata
        }